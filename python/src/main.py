import logging
from typing import Dict, Tuple, List

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string
from ruamel.yaml import YAML, CommentedMap
from os.path import exists

from argparse import ArgumentParser, Namespace

LOG = logging.getLogger(__name__)


def copy_data(workbook: Workbook, sport, sport_columns, sheet_row, row):
    LOG.info(f"Copying data for {row[2]} {row[3]} to {sport['tab']}")
    sport_tab_name = sport["tab"]
    destination_sheet = workbook[sport_tab_name]
    for column in sport_columns:
        if "from" in column:
            from_column = column["from"]
            to_column = column["to"]
            if from_column == "tab name":
                value = sport_tab_name
            else:
                value = row[column_index_from_string(from_column) - 1]

            destination_sheet[f"{to_column}{sheet_row[sport_tab_name]}"] = value

    sheet_row[sport_tab_name] += 1


def copy_row(
    yaml_config: CommentedMap,
    workbook: Workbook,
    sheet_row: Dict[str, int],
    row: List[str],
) -> bool:
    if row[2] == "":
        # No first name, so skip
        return False

    # Find the preferred sport
    for sport in yaml_config["sports"]:
        column_number = column_index_from_string(sport["school column"])
        if row[column_number - 1] == 1:
            # Found the sport
            copy_data(workbook, sport, yaml_config["by sport columns"], sheet_row, row)
            return True

    # If we are here we didn't find the sport
    LOG.error(f"Could not find sport for {row[0]} {row[1]} {row[2]} {row[3]}")
    return False


def trim_row_data(row_in: Tuple) -> List:
    row_out = []
    for cell in row_in:
        if cell.value is None or cell.value == "#VALUE!":
            row_out.append("")
        elif isinstance(cell.value, str):
            row_out.append(cell.value.strip())
        else:
            row_out.append(cell.value)

    return row_out


def process_school(
    yaml_config: CommentedMap,
    workbook_in: Workbook,
    workbook_out: Workbook,
    school: CommentedMap,
    sheet_row: Dict[str, int],
):
    LOG.info(f"Processing school {school['name']}")
    work_sheet = workbook_in[school["tab"]]
    reading_heading = True
    count = 0
    for row in work_sheet.rows:
        trimmed_row = trim_row_data(row)
        if reading_heading:
            if (
                trimmed_row[0] == "No"
                and trimmed_row[1] == "School"
                and trimmed_row[2] == "First name"
                and trimmed_row[3] == "Surname"
            ):
                reading_heading = False
                LOG.info(f"Found heading row for {school['name']}")

        elif copy_row(yaml_config, workbook_out, sheet_row, trimmed_row):
            count += 1

    LOG.info(f"Processed {count} rows for {school['name']}")


def fill_sheets(
    yaml_config: CommentedMap, workbook_in: Workbook, workbook_out: Workbook
):
    sheet_row = {sport["tab"]: 2 for sport in yaml_config["sports"]}
    for school in yaml_config["schools"]:
        # Process school
        process_school(yaml_config, workbook_in, workbook_out, school, sheet_row)


def format_worksheet(worksheet):
    # Freeze the top row
    worksheet.freeze_panes = "A2"

    # Set row height
    worksheet.row_dimensions[1].height = 45


def add_sheets(yaml_config: CommentedMap, workbook: Workbook):
    # Added the sheets
    for sheet in yaml_config["sports"]:
        LOG.info(f"Create sheet for {sheet['tab']}")
        tab_name = sheet["tab"]
        worksheet = workbook.create_sheet(tab_name)

        # Add the columns
        for column in yaml_config["by sport columns"]:
            worksheet[f"{column['to']}1"] = column["name"]
            worksheet[f"{column['to']}1"].alignment = Alignment(
                vertical="top", horizontal="center", wrap_text=True
            )
            worksheet.column_dimensions[column["to"]].width = column["width"]

        format_worksheet(worksheet)


def create_blank_workbook(yaml_config: CommentedMap) -> Workbook:
    workbook = Workbook()
    add_sheets(yaml_config, workbook)

    # Remove the default sheet
    del workbook["Sheet"]

    return workbook


def find_row(row, sheet_in):
    for row_in in sheet_in.rows:
        # Trim the row
        trimmed_row = trim_row_data(row_in)
        if (
            trimmed_row[1] == row[1]
            and trimmed_row[2] == row[2]
            and trimmed_row[3] == row[3]
            and trimmed_row[4] == row[4]
        ):
            # Found the row
            return trimmed_row

    return None


def delete_rows(yaml_config, workbook_in, workbook_out):
    for sport in yaml_config["sports"]:
        sheet_in = workbook_in[sport["tab"]]
        sheet_out = workbook_out[sport["tab"]]

        rows_to_delete = []

        for row_number, row in enumerate(sheet_out.rows, start=1):
            if row_number == 1:
                # Skip the heading row
                continue

            trimmed_row_out = trim_row_data(row)

            # Does this row exist in the new file
            if find_row(trimmed_row_out, sheet_in) is None:
                LOG.info(
                    f"Adding row to delete {row_number} "
                    f"{trimmed_row_out[1]} {trimmed_row_out[2]} {trimmed_row_out[3]} "
                    f"from {sport['tab']}"
                )
                rows_to_delete.append(row_number)

        # Delete the rows
        for row_number in reversed(rows_to_delete):
            LOG.info(f"Deleting row {row_number} from {sport['tab']}")
            sheet_out.delete_rows(row_number)


def add_rows(yaml_config, workbook_in, workbook_out):
    for sport in yaml_config["sports"]:
        sheet_in = workbook_in[sport["tab"]]
        sheet_out = workbook_out[sport["tab"]]

        rows_to_add = []

        for row_number, row in enumerate(sheet_in.rows, start=1):
            if row_number == 1:
                # Skip the heading row
                continue

            trimmed_row_in = trim_row_data(row)

            # Does this row exist in the new file
            if find_row(trimmed_row_in, sheet_out) is None:
                rows_to_add.append(row)

        # Delete the rows
        for row in rows_to_add:
            sheet_out.add_row(row)


def update_existing_rows(yaml_config, workbook_in, workbook_out):
    for sport in yaml_config["sports"]:
        sheet_in = workbook_in[sport["tab"]]
        sheet_out = workbook_out[sport["tab"]]

        for row_number, row in enumerate(sheet_out.rows, start=1):
            if row_number == 1:
                # Skip the heading row
                continue

            trimmed_row_out = trim_row_data(row)

            # Does this row exist in the both files
            row_in = find_row(trimmed_row_out, sheet_in)
            if row_in is not None:
                # Do the updates
                for column_index in range(4, 16):
                    sheet_out.cell(row_number, column_index + 1).value = row_in[
                        column_index
                    ]


def update_workbook(yaml_config, workbook_in, workbook_out):
    # Pass 1: Delete rows that are not needed
    delete_rows(yaml_config, workbook_in, workbook_out)

    # Pass 2: Add new rows
    add_rows(yaml_config, workbook_in, workbook_out)

    # Pass 3: Update exist
    update_existing_rows(yaml_config, workbook_in, workbook_out)


def do_copy(args: Namespace):
    # Load the YAML
    with open(args.yaml_file, "r") as yaml_file:
        yaml = YAML()
        yaml_config = yaml.load(yaml_file)

    # Load the input Excel file
    workbook_in = load_workbook(args.input_file, read_only=True, data_only=True)

    # If the output file exists read it
    if exists(args.output_file):
        LOG.info(f"Updating file {args.output_file}")

        # Create a temp workbook with the new values
        LOG.info("Creating a temporary workbook")
        workbook_temp = create_blank_workbook(yaml_config)
        fill_sheets(yaml_config, workbook_in, workbook_temp)

        # Load the current output file
        workbook_out = load_workbook(args.output_file)

        # Start writing the output file
        update_workbook(
            yaml_config,
            workbook_temp,
            workbook_out,
        )

    else:
        LOG.info(f"Creating file {args.output_file}")
        workbook_out = create_blank_workbook(yaml_config)
        fill_sheets(yaml_config, workbook_in, workbook_out)

    # Write out the Excel file
    workbook_out.save(f"{args.output_file}")


def user_confirm(question: str):
    reply = str(input(f"{question} (y/n): ")).lower().strip()
    return reply[0] in ["y", "yes"]


def main(args: Namespace):
    if not exists(args.input_file):
        LOG.error(f"File {args.input_file} does not exist")
        return

    if not exists(args.yaml_file):
        LOG.error(f"File {args.yaml_file} does not exists")
        return

    do_copy(args)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)

    parser = ArgumentParser("Copy the data across")

    parser.add_argument("input_file", help="The input Excel to parse")
    parser.add_argument("output_file", help="The output Excel file")
    parser.add_argument("yaml_file", help="The yaml configuration file")
    main(parser.parse_args())
